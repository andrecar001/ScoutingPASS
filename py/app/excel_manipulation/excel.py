from openpyxl import load_workbook
import os
import information
def isDouble(value):
    try:
        float_value = float(value)
        return True
    except ValueError:
        return False

# print(wb.sheetnames)
def populateSheet(workbookPath, matchDataPath, pitDataPath):
    path = 'py/app/data/Test.xlsx'
    workbookPath = os.path.abspath(workbookPath)
    # wb = load_workbook(filename=workbookPath, read_only=False, keep_vba=True)
    wb = load_workbook(workbookPath)
    filePath = 'py/app/data/match_scouting_data.txt'

    match_sheet_name = 'ScoutingInfo'
    pit_sheet_name = 'PitScoutingInfo'
    match_sheet = wb[match_sheet_name]
    pit_sheet = wb[pit_sheet_name]
    match_lines = information.getAllMatchInfoList(matchDataPath)
    pit_lines = information.getAllPitInfoList(pitDataPath)
    # Update Match Sheet
    for row_idx, row in enumerate(match_lines, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = match_sheet.cell(row=row_idx+1,column=col_idx)
            if isDouble(value):
                cell.value = float(value)
                if '.' in value:
                    cell.number_format = '0.0'
                else:
                    cell.number_format = '0'
            else:
                cell.value = value
        
    for row_idx, row in enumerate(pit_lines, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = pit_sheet.cell(row=row_idx+1,column=col_idx)
            if isDouble(value):
                cell.value = float(value)
                if '.' in value:
                    cell.number_format = '0.0'
                else:
                    cell.number_format = '0'
            else:
                cell.value = value
    wb.save(workbookPath)
    wb.close()
    return
